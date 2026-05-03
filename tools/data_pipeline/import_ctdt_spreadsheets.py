from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import io
from pathlib import Path
import re
import sys
import unicodedata
from typing import Iterable, Optional

import requests


def _find_repo_root() -> Path:
    for parent in Path(__file__).resolve().parents:
        if (parent / "main.py").exists() and (parent / "services").is_dir():
            return parent
    return Path(__file__).resolve().parents[2]


PROJECT_ROOT = _find_repo_root()
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config.settings import settings
from services.document_service import reingest_uploaded_documents


DEFAULT_SOURCE_MD = (
    settings.QA_CORPUS_ROOT
    / "Sổ tay sinh viên các năm"
    / "6. Sổ tay sinh viên các năm"
    / "8. SO TAY SINH VIEN 2025-2026.md"
)
DEFAULT_OUTPUT_DIR = settings.RAG_UPLOAD_ROOT / "student_handbook_rag"

GOOGLE_SHEET_RE = re.compile(
    r"https://docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]+)/edit\?gid=([0-9]+)",
    flags=re.IGNORECASE,
)
TAI_VE_LINE_RE = re.compile(
    r"^\s*(?P<label>.+?)\s+T(?:ả|a)i\s+v(?:ề|e)\s*:\s*$",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class SheetSpec:
    label: str
    spreadsheet_id: str
    gid: str
    source_url: str


def _safe_print(message: str = "", *, error: bool = False) -> None:
    stream = sys.stderr if error else sys.stdout
    encoding = stream.encoding or "utf-8"
    try:
        print(message, file=stream)
    except UnicodeEncodeError:
        fallback = message.encode(encoding, errors="backslashreplace").decode(encoding)
        print(fallback, file=stream)


def _normalize_ascii(value: str) -> str:
    folded = unicodedata.normalize("NFKD", value.casefold())
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    folded = folded.replace("đ", "d")
    return re.sub(r"\s+", " ", folded).strip()


def _slugify(value: str) -> str:
    folded = (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    return re.sub(r"[^a-z0-9]+", "_", folded).strip("_") or "sheet"


def _collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def _parse_google_sheet_url(raw_line: str) -> Optional[tuple[str, str, str]]:
    compact = re.sub(r"\s+", "", raw_line)
    match = GOOGLE_SHEET_RE.search(compact)
    if not match:
        return None
    spreadsheet_id = match.group(1)
    gid = match.group(2)
    canonical = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit?gid={gid}"
    return spreadsheet_id, gid, canonical


def extract_sheet_specs_from_markdown(path: Path) -> list[SheetSpec]:
    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay file markdown nguon: {path}")

    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    specs: list[SheetSpec] = []
    current_label: Optional[str] = None

    for line in lines:
        collapsed = _collapse_spaces(line)
        label_match = TAI_VE_LINE_RE.match(collapsed)
        if label_match:
            current_label = label_match.group("label").strip(" -:")
            continue

        parsed = _parse_google_sheet_url(line)
        if parsed is None:
            continue

        spreadsheet_id, gid, canonical = parsed
        label = current_label or f"Sheet {len(specs) + 1}"
        specs.append(
            SheetSpec(
                label=label,
                spreadsheet_id=spreadsheet_id,
                gid=gid,
                source_url=canonical,
            )
        )

    unique: list[SheetSpec] = []
    seen: set[tuple[str, str]] = set()
    for spec in specs:
        key = (spec.spreadsheet_id, spec.gid)
        if key in seen:
            continue
        seen.add(key)
        unique.append(spec)
    return unique


def _parse_csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows: list[list[str]] = []
    for row in reader:
        cleaned = [re.sub(r"\s+", " ", str(cell).strip()) for cell in row]
        while cleaned and not cleaned[-1]:
            cleaned.pop()
        if any(cleaned):
            rows.append(cleaned)
    return rows


def _to_markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_Khong co du lieu dang bang._"

    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]

    def esc(cell: str) -> str:
        return cell.replace("|", "\\|")

    header = normalized[0]
    body = normalized[1:] or [[""] * width]
    lines = [
        "| " + " | ".join(esc(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(esc(cell) for cell in row) + " |")
    return "\n".join(lines)


def _download_google_sheet_csv(spec: SheetSpec, *, timeout: int = 45) -> list[list[str]]:
    export_url = (
        f"https://docs.google.com/spreadsheets/d/{spec.spreadsheet_id}/export"
        f"?format=csv&gid={spec.gid}"
    )
    response = requests.get(export_url, timeout=timeout)
    response.raise_for_status()
    return _parse_csv_rows(response.content)


def _load_xlsx_rows(path: Path, *, sheet_name: Optional[str] = None) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Chua co openpyxl trong moi truong. Cai dat openpyxl hoac dung Google Sheets URL."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    outputs: list[tuple[str, list[list[str]]]] = []
    worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
    for ws in worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [re.sub(r"\s+", " ", str(cell).strip()) if cell is not None else "" for cell in row]
            while cleaned and not cleaned[-1]:
                cleaned.pop()
            if any(cleaned):
                rows.append(cleaned)
        outputs.append((ws.title, rows))
    return outputs


def _build_markdown_document(
    *,
    title: str,
    cohort_tag: str,
    academic_year: str,
    source_label: str,
    source_url: str,
    rows: list[list[str]],
) -> str:
    generated_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    table = _to_markdown_table(rows)
    row_count = len(rows)
    col_count = max((len(row) for row in rows), default=0)
    return (
        "---\n"
        f'title: "{title}"\n'
        'document_type: "ctdt_sheet_import"\n'
        f'cohort_tag: "{cohort_tag}"\n'
        f'academic_year: "{academic_year}"\n'
        f'source_label: "{source_label}"\n'
        f'source_url: "{source_url}"\n'
        f'generated_at: "{generated_at}"\n'
        f"row_count: {row_count}\n"
        f"column_count: {col_count}\n"
        "---\n\n"
        f"# {title}\n\n"
        f"- Cohort: `{cohort_tag}`\n"
        f"- Academic year: `{academic_year}`\n"
        f"- Source label: `{source_label}`\n"
        f"- Source URL: {source_url}\n\n"
        "## Program Data\n\n"
        f"{table}\n"
    )


def _write_markdown(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Import CTDT spreadsheets (Google Sheets or local XLSX) to markdown files under "
            "data/rag_uploads/student_handbook_rag and optionally re-ingest vector store."
        )
    )
    parser.add_argument(
        "--source-md",
        type=Path,
        default=DEFAULT_SOURCE_MD,
        help="Markdown file that contains Google Sheets links (default: SO TAY 2025-2026).",
    )
    parser.add_argument(
        "--cohort-tag",
        type=str,
        default="k24",
        help="Cohort tag for generated markdown metadata (e.g. k24, k23).",
    )
    parser.add_argument(
        "--academic-year",
        type=str,
        default="2025-2026",
        help="Academic year metadata for generated markdown files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory to store generated markdown files.",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        action="append",
        default=[],
        help="Optional local XLSX path(s) to import in addition to Google Sheets.",
    )
    parser.add_argument(
        "--xlsx-sheet",
        type=str,
        default=None,
        help="Optional sheet name when importing local XLSX files.",
    )
    parser.add_argument(
        "--extra-sheet-url",
        type=str,
        action="append",
        default=[],
        help="Optional extra Google Sheets URL(s).",
    )
    parser.add_argument(
        "--skip-reingest",
        action="store_true",
        help="Only generate markdown files, do not run re-ingest.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview discovered sources and target filenames without writing or re-ingesting.",
    )
    return parser.parse_args()


def _sheet_spec_from_url(url: str, fallback_label: str) -> SheetSpec:
    parsed = _parse_google_sheet_url(url)
    if parsed is None:
        raise ValueError(f"Google Sheets URL khong hop le: {url}")
    spreadsheet_id, gid, canonical = parsed
    return SheetSpec(
        label=fallback_label,
        spreadsheet_id=spreadsheet_id,
        gid=gid,
        source_url=canonical,
    )


def _build_output_filename(
    *,
    cohort_tag: str,
    academic_year: str,
    index: int,
    label: str,
) -> str:
    return (
        f"ctdt_{_slugify(cohort_tag)}_{_slugify(academic_year)}_"
        f"{index:02d}_{_slugify(label)}.md"
    )


def _iter_google_specs(args: argparse.Namespace) -> list[SheetSpec]:
    specs = extract_sheet_specs_from_markdown(args.source_md.resolve())
    for idx, url in enumerate(args.extra_sheet_url, start=1):
        specs.append(_sheet_spec_from_url(url, fallback_label=f"extra_sheet_{idx}"))
    return specs


def main() -> int:
    args = parse_args()
    cohort_tag = _normalize_ascii(args.cohort_tag) or "k24"
    academic_year = _collapse_spaces(args.academic_year) or "2025-2026"
    output_dir = args.output_dir.resolve()

    google_specs = _iter_google_specs(args)
    xlsx_files = [path.resolve() for path in args.xlsx]

    if not google_specs and not xlsx_files:
        raise RuntimeError("Khong tim thay nguon Google Sheets/XLSX nao de import.")

    generated_files: list[Path] = []
    row_summaries: list[str] = []
    index = 1

    for spec in google_specs:
        rows = _download_google_sheet_csv(spec)
        title = f"CTDT {cohort_tag.upper()} - {spec.label}"
        filename = _build_output_filename(
            cohort_tag=cohort_tag,
            academic_year=academic_year,
            index=index,
            label=spec.label,
        )
        destination = output_dir / filename
        markdown = _build_markdown_document(
            title=title,
            cohort_tag=cohort_tag,
            academic_year=academic_year,
            source_label=spec.label,
            source_url=spec.source_url,
            rows=rows,
        )
        if not args.dry_run:
            _write_markdown(destination, markdown)
        generated_files.append(destination)
        row_summaries.append(f"{filename} (rows={len(rows)})")
        index += 1

    for xlsx_path in xlsx_files:
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Khong tim thay XLSX: {xlsx_path}")
        for sheet_title, rows in _load_xlsx_rows(xlsx_path, sheet_name=args.xlsx_sheet):
            label = f"{xlsx_path.stem} - {sheet_title}"
            title = f"CTDT {cohort_tag.upper()} - {label}"
            filename = _build_output_filename(
                cohort_tag=cohort_tag,
                academic_year=academic_year,
                index=index,
                label=label,
            )
            destination = output_dir / filename
            markdown = _build_markdown_document(
                title=title,
                cohort_tag=cohort_tag,
                academic_year=academic_year,
                source_label=label,
                source_url=str(xlsx_path),
                rows=rows,
            )
            if not args.dry_run:
                _write_markdown(destination, markdown)
            generated_files.append(destination)
            row_summaries.append(f"{filename} (rows={len(rows)})")
            index += 1

    _safe_print(f"Generated {len(generated_files)} markdown file(s) into {output_dir}")
    for summary in row_summaries:
        _safe_print(f"- {summary}")

    if args.dry_run or args.skip_reingest:
        if args.dry_run:
            _safe_print("Dry run: skip writing/re-ingest.")
        else:
            _safe_print("Skip re-ingest theo tham so --skip-reingest.")
        return 0

    total_files, total_chunks = reingest_uploaded_documents()
    _safe_print(
        f"Re-ingest completed: reloaded {total_files} files, added {total_chunks} chunk(s) in vector store."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
